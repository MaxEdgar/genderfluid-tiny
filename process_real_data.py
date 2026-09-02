#!/usr/bin/env python3
"""
Process real SSA baby names and Census 2020 first names data
into the training format for genderfluid-tiny.

Data sources:
- SSA national baby names (1880-2020): github.com/hackerb9/ssa-baby-names
- Census 2020 first names by sex: census.gov
"""

import json
import os
from collections import defaultdict


def process_ssa_data(filepath):
    """
    Process SSA alldata.txt format: Name,Sex,Count,Year
    Returns dict: name -> {"female": total_count, "male": total_count}
    """
    name_counts = defaultdict(lambda: {"female": 0, "male": 0})

    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split(",")
            if len(parts) != 4:
                continue
            name, sex, count_str, year = parts
            try:
                count = int(count_str)
            except ValueError:
                continue

            name = name.strip()
            sex = sex.strip().upper()

            if sex == "F":
                name_counts[name]["female"] += count
            elif sex == "M":
                name_counts[name]["male"] += count

    return dict(name_counts)


def process_census_data(filepath):
    """
    Process Census 2020 xlsx: Name, Rank, Count, Proportion, Cumulative, Male, Female
    Returns dict: name -> {"female": count, "male": count}
    """
    import openpyxl

    name_counts = {}

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Skip header rows (3 rows of headers)
    rows = list(ws.iter_rows(min_row=4, values_only=True))

    for row in rows:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        male_count = row[5] if row[5] else 0
        female_count = row[6] if row[6] else 0

        try:
            male_count = max(0, int(male_count))
            female_count = max(0, int(female_count))
        except (ValueError, TypeError):
            continue

        if name:
            name_counts[name] = {"female": female_count, "male": male_count}

    wb.close()
    return name_counts


def combine_datasets(ssa_counts, census_counts):
    """
    Combine SSA and Census counts.
    Census data is more recent (2020) and gives a snapshot.
    SSA data gives historical totals.
    We weight Census data more heavily (3x) since it's a more recent snapshot.
    """
    all_names = set(ssa_counts.keys()) | set(census_counts.keys())
    combined = {}

    for name in all_names:
        ssa = ssa_counts.get(name, {"female": 0, "male": 0})
        cen = census_counts.get(name, {"female": 0, "male": 0})

        # Weight Census data more heavily
        female = ssa["female"] + cen["female"] * 3
        male = ssa["male"] + cen["male"] * 3

        combined[name] = {"female": female, "male": male}

    return combined


def label_name(female_count, male_count):
    """
    Label a name based on statistical association.
    Uses a ratio threshold to determine label.
    """
    total = female_count + male_count

    if total == 0:
        return None, 0.0

    female_ratio = female_count / total
    male_ratio = male_count / total

    # Use 85% threshold for strong association
    if female_ratio >= 0.85:
        return "girl-associated", female_ratio
    elif male_ratio >= 0.85:
        return "boy-associated", male_ratio
    else:
        return "uncertain", max(female_ratio, male_ratio)


def create_training_data(combined_counts, min_count=50):
    """
    Create training entries from combined counts.
    Only includes names with sufficient total count.
    """
    entries = []
    skipped = 0

    for name, counts in combined_counts.items():
        total = counts["female"] + counts["male"]
        if total < min_count:
            skipped += 1
            continue

        label, confidence = label_name(counts["female"], counts["male"])
        if label is None:
            continue

        # Weight by log of total count (more common names have more weight)
        import math
        weight = min(1.0, math.log10(total + 1) / 7.0)

        entries.append({
            "name": name,
            "label": label,
            "weight": round(weight, 3),
        })

    return entries, skipped


def main():
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")

    ssa_path = os.path.join(data_dir, "ssa_alldata.txt")
    census_path = os.path.join(data_dir, "census_2020_firstnames_sex.xlsx")
    output_path = os.path.join(data_dir, "names.jsonl")

    # Check files exist
    if not os.path.exists(ssa_path):
        print(f"SSA data not found: {ssa_path}")
        print("Download from: https://raw.githubusercontent.com/hackerb9/ssa-baby-names/main/alldata.txt")
        return
    if not os.path.exists(census_path):
        print(f"Census data not found: {census_path}")
        print("Download from: https://www2.census.gov/topics/genealogy/2020surnames/Names2020_FirstNames_Sex.xlsx")
        return

    print("Processing SSA data...")
    ssa_counts = process_ssa_data(ssa_path)
    print(f"  SSA unique names: {len(ssa_counts):,}")

    print("Processing Census data...")
    census_counts = process_census_data(census_path)
    print(f"  Census unique names: {len(census_counts):,}")

    print("Combining datasets...")
    combined = combine_datasets(ssa_counts, census_counts)
    print(f"  Combined unique names: {len(combined):,}")

    print("Creating training entries...")
    entries, skipped = create_training_data(combined, min_count=50)
    print(f"  Entries with sufficient data: {len(entries):,}")
    print(f"  Skipped (insufficient count): {skipped:,}")

    # Count labels
    girl_count = sum(1 for e in entries if e["label"] == "girl-associated")
    boy_count = sum(1 for e in entries if e["label"] == "boy-associated")
    uncertain_count = sum(1 for e in entries if e["label"] == "uncertain")

    print(f"\nDataset statistics")
    print(f"-" * 40)
    print(f"Total entries: {len(entries):,}")
    print(f"Girl-associated: {girl_count:,}")
    print(f"Boy-associated: {boy_count:,}")
    print(f"Uncertain: {uncertain_count:,}")
    print(f"Unique names: {len(entries):,}")

    # Write to file
    with open(output_path, "w", encoding="utf-8") as f:
        for entry in entries:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")

    print(f"\nSaved to: {output_path}")


if __name__ == "__main__":
    main()
