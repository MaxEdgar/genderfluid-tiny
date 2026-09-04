#!/usr/bin/env python3
"""Build a multinational name-gender association dataset from official sources.

Sources:
  America:
    - U.S. Social Security Administration baby names (data/ssa_alldata.txt)
    - U.S. Census Bureau 2020 first names (data/census_2020_firstnames_sex.xlsx)
  Europe:
    - INSEE France "Base prénoms" 1900-2024 (downloaded from data.gouv.fr)
    - INE Spain "Nombres y apellidos más frecuentes" Padrón 2025 (given
      names with >=20 holders, by sex, official xlsx)
    - Institute for Language and Folklore (Sweden) newborn top-name tables
      2023-2025 with real counts (official xlsx)
    - PESEL register, Poland (full list of living people's first names
      with counts, official Ministry of Digitisation CSV)
  Asia:
    - Meiji Yasuda Life newborn name survey, Japan (downloaded JSON)
    - Ministry of Public Security annual name reports, China (official
      statistics embedded below, published via state media)

Each source contributes per-name counts by sex. Names are merged across
sources and labeled by statistical association: >=85% of recorded births
as female -> girl-associated, >=85% male -> boy-associated, otherwise
uncertain. Rows are weighted by frequency so common names dominate.

Outputs data/names.jsonl plus 80/10/10 train/validation/test splits that
are partitioned by normalized name (a name never appears in two splits).

Japan and China publish associations rather than raw counts, so those
sources use synthetic counts that preserve the reported association
(100 for single-gender, 50/50 for mixed). Other Asian and African
countries either do not publish gendered name statistics or block
automated downloads; this script is structured so additional sources can
be added as functions that return {name: (girl_count, boy_count)}.
"""

import hashlib
import json
import os
import sys
import urllib.request

_UA = {"User-Agent": "Mozilla/5.0 (genderfluid-tiny data fetch; research use)"}

from genderfluid.preprocessing import normalize_name

ASSOC_THRESHOLD = 0.85
TRAIN_FRACTION = 0.80
VAL_FRACTION = 0.10

INSEE_DATASET_API = "https://www.data.gouv.fr/api/1/datasets/base-prenoms-2024-insee/"
SSA_ZIP_URL = "https://www.ssa.gov/oact/babynames/names.zip"
MEIJI_INDEX_URL = "https://www.meijiyasuda.co.jp/enjoy/ranking/assets/json/index_name.json"

# Spain: INE "Nombres y apellidos mas frecuentes" - Padron-based frequency of
# every given name with >=20 holders nationwide, by sex (Hombres/Mujeres
# sheets: rank, name, frequency, mean age). Official INE publication.
ES_NAMES_URL = "https://www.ine.es/daco/daco42/nombyapel/nombres_por_edad_media.xlsx"

# Sweden: Institute for Language and Folklore (official, skatteverket data)
# newborn top-name tables with real counts. One file per year.
SE_ISOF_FILES = [
    (2025, "https://www.isof.se/download/18.331a790519cc9cca613dd136/1773161563268/Statistik%20babynamn%202025%20ny.xlsx"),
    (2024, "https://www.isof.se/download/18.359aef78194d854194abcde2/1740489507331/Statistik%20babynamn%202024.xlsx"),
    (2023, "https://www.isof.se/download/18.7ada16dc193deb67224b7be6/1736421128822/Tilltalsnamn%20nyf%C3%B6dda%202023.xlsx"),
]

# Poland: full PESEL register snapshot (19.01.2024) - every first name of
# living people with occurrence counts, by legal sex. Published by the
# Ministry of Digitisation on the official open-data portal (dane.gov.pl,
# dataset 1667 "Lista imion wystepujacych w rejestrze PESEL"). Names with
# single occurrences and deceased persons are excluded by the publisher.
PL_PESEL_FILES = [
    ("male", "https://api.dane.gov.pl/media/resources/20240126/8_-_Wykaz_imion_m%C4%99skich_os%C3%B3b_%C5%BCyj%C4%85cych_wg_pola_imi%C4%99_pierwsze_wyst%C4%99puj%C4%85cych_w_rejestrze_PESEL_bez_zgon%C3%B3w.csv"),
    ("female", "https://api.dane.gov.pl/media/resources/20240126/8_-_Wykaz_imion_%C5%BCe%C5%84skich_os%C3%B3b_%C5%BCyj%C4%85cych_wg_pola_imi%C4%99_pierwsze_wyst%C4%99puj%C4%85cych_w_rejestrze_PESEL_bez_zgon%C3%B3w.csv"),
]

# Official statistics from the Ministry of Public Security (China) annual
# name reports, published via state media. Facts (name -> sex association)
# from official government statistics; sources cited in data/README.md.
CHINA_MPS_NAMES = {
    # 2021 report: male newborns, top 10 (People's Daily, 2022-01-24)
    "boy-associated": [
        "沐宸", "浩宇", "沐辰", "茗泽", "奕辰", "宇泽", "浩然", "奕泽", "宇轩", "沐阳",
        # 2020 report: male newborns, top 10 (China News, 2021-02-08)
        "亦辰", "宇辰", "子墨", "宇航", "梓豪", "亦宸",
        # 2019 report: male newborns, top 10 (Xinhua, 2020-01-21)
        "铭泽", "子睿", "子轩", "梓睿",
        # 2018 report: male newborns, top 10 (China Daily, 2019-01-30)
        "浩轩", "子豪",
        # 2021 report: single-char names used more by males
        "伟", "杰", "勇", "涛", "军", "强",
        # 2021 report: double-char names used more by males
        "建华",
    ],
    "girl-associated": [
        # 2021 report: female newborns, top 10
        "若汐", "一诺", "艺涵", "依诺", "梓涵", "苡沫", "雨桐", "欣怡", "语桐", "语汐",
        # 2020 report: female newborns, top 10
        "欣妍", "可欣", "梦瑶",
        # 2019 report: female newborns, top 10 (Xinhua, 2020-01-21)
        "诗涵", "佳怡", "佳琪", "梓萱",
        # 2018 report: female newborns, top 10 (China Daily, 2019-01-30)
        "可馨",
        # 2021 report: single-char names used more by females
        "敏", "静", "丽", "艳",
        # 2021 report: double-char names used more by females
        "秀英", "桂英", "秀兰", "玉兰", "婷婷", "桂兰", "玉梅", "秀珍", "海燕",
    ],
}


def _load_ssa(path: str, url: str = None) -> dict:
    """
    Parse SSA national baby names.

    Prefers the official names.zip (all years, names with 5+ occurrences,
    roughly a million unique spellings). Falls back to a local aggregated
    file if the download fails.
    """
    base = os.path.dirname(os.path.abspath(__file__))
    cache_zip = os.path.join(base, "data", "raw", "ssa_names.zip")

    if url:
        try:
            import zipfile
            if not os.path.exists(cache_zip) or os.path.getsize(cache_zip) == 0:
                print("  Downloading SSA national data (names.zip)...")
                req = urllib.request.Request(url, headers={
                    "User-Agent": ("Mozilla/5.0 (X11; Linux x86_64) "
                                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                                    "Chrome/126.0 Safari/537.36"),
                    "Accept": "*/*",
                })
                with urllib.request.urlopen(req, timeout=600) as resp, \
                        open(cache_zip, "wb") as out:
                    while True:
                        chunk = resp.read(1024 * 1024)
                        if not chunk:
                            break
                        out.write(chunk)
            counts: dict = {}
            with zipfile.ZipFile(cache_zip) as zf:
                for member in zf.namelist():
                    if not member.startswith("yob") or not member.endswith(".txt"):
                        continue
                    for line in zf.read(member).decode("utf-8", "replace").splitlines():
                        parts = line.split(",")
                        if len(parts) < 3:
                            continue
                        name = normalize_name(parts[0])
                        sex = parts[1].strip().upper()
                        try:
                            count = int(parts[2])
                        except ValueError:
                            continue
                        if not name or sex not in ("F", "M"):
                            continue
                        bucket = counts.setdefault(name, [0, 0])
                        bucket[0 if sex == "F" else 1] += count
            print(f"  SSA names.zip: {len(counts):,} unique names")
            return counts
        except Exception as exc:
            print(f"  WARNING: SSA names.zip failed ({exc}); using local file")

    counts: dict = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split(",")
            if len(parts) < 3:
                continue
            name = normalize_name(parts[0])
            sex = parts[1].strip().upper()
            try:
                count = int(parts[2])
            except ValueError:
                continue
            if not name or sex not in ("F", "M"):
                continue
            bucket = counts.setdefault(name, [0, 0])
            bucket[0 if sex == "F" else 1] += count
    return counts


def _load_census_xlsx(path: str) -> dict:
    """Parse Census 2020 first names workbook: FIRST NAME, MALE, FEMALE."""
    try:
        import openpyxl
    except ImportError:
        print("WARNING: openpyxl not installed, skipping Census source")
        return {}
    counts: dict = {}
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3 or row is None or len(row) < 7:
            continue
        name = normalize_name(str(row[0]))
        if not name:
            continue
        try:
            male = int(row[5] or 0)
            female = int(row[6] or 0)
        except (ValueError, TypeError):
            continue
        if male + female == 0:
            continue
        bucket = counts.setdefault(name, [0, 0])
        bucket[0] += female
        bucket[1] += male
    wb.close()
    return counts


def _load_insee() -> dict:
    """Download the INSEE first-names parquet and aggregate counts by name."""
    try:
        import pyarrow.parquet as pq
    except ImportError:
        print("ERROR: pyarrow is required for the INSEE source. "
              "Install it with: pip install pyarrow")
        sys.exit(1)

    print("Resolving INSEE dataset resource...")
    with urllib.request.urlopen(INSEE_DATASET_API, timeout=60) as resp:
        meta = json.loads(resp.read().decode("utf-8"))
    resources = meta.get("resources", [])
    if not resources:
        print("ERROR: no resources found for INSEE dataset")
        sys.exit(1)
    # Prefer the newest parquet resource (the API returns them sorted).
    url = resources[0]["url"]
    print(f"  Downloading: {url}")

    local_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "data", "raw", "insee_prenoms.parquet")
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    if not os.path.exists(local_path) or os.path.getsize(local_path) == 0:
        for attempt in range(3):
            try:
                with urllib.request.urlopen(url, timeout=600) as resp, \
                        open(local_path, "wb") as out:
                    total = int(resp.headers.get("Content-Length", 0) or 0)
                    done = 0
                    while True:
                        chunk = resp.read(1024 * 1024)
                        if not chunk:
                            break
                        out.write(chunk)
                        done += len(chunk)
                        if total:
                            print(f"\r  Downloaded {done/1048576:.1f} MB / {total/1048576:.1f} MB", end="")
                print()
            except Exception as exc:
                print(f"\n  Download attempt {attempt + 1} failed: {exc}")
                continue
            if total and os.path.getsize(local_path) < total:
                print(f"  Incomplete download ({os.path.getsize(local_path)} of {total} bytes), retrying...")
                continue
            break
        if os.path.getsize(local_path) == 0:
            print("ERROR: could not download INSEE dataset")
            sys.exit(1)

    print("  Aggregating records...")
    counts: dict = {}
    pf = pq.ParquetFile(local_path)
    for batch in pf.iter_batches(batch_size=100_000):
        cols = batch.to_pydict()
        sexe = cols.get("sexe") or cols.get("SEXE")
        preusuel = (cols.get("preusuel") or cols.get("prenom")
                    or cols.get("PREUSUEL"))
        nombre = (cols.get("nombre") or cols.get("valeur")
                  or cols.get("NOMBRE"))
        if sexe is None or preusuel is None or nombre is None:
            print("ERROR: unexpected INSEE columns:", list(cols.keys()))
            sys.exit(1)
        for s, p, n in zip(sexe, preusuel, nombre):
            name = normalize_name(str(p))
            if not name or s is None:
                continue
            try:
                c = int(n)
                sex = int(s)
            except (ValueError, TypeError):
                continue
            # INSEE: sexe 1 = male, 2 = female
            bucket = counts.setdefault(name, [0, 0])
            if sex == 2:
                bucket[0] += c
            elif sex == 1:
                bucket[1] += c
    return counts


def _load_meiji_yasuda() -> dict:
    """
    Download Japan's Meiji Yasuda newborn-name survey index.

    Meiji Yasuda Life publishes an annual survey of the most popular
    newborn names (kanji), each labeled male/female/mixed. This is the
    standard public source for Japanese name-gender associations.

    Returns {name: (girl_count, boy_count)} with synthetic counts that
    preserve the survey's gender association (no raw counts are public).
    """
    print("Downloading Japan Meiji Yasuda name survey...")
    req = urllib.request.Request(MEIJI_INDEX_URL, headers={
        "User-Agent": ("Mozilla/5.0 (X11; Linux x86_64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/126.0 Safari/537.36"),
        "Accept": "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception as exc:
        print(f"  WARNING: Meiji Yasuda download failed ({exc}); "
              "skipping Japan source")
        return {}

    # Synthetic counts preserve the reported association. The survey does
    # not publish raw counts; confident single-gender names get 100, mixed
    # get 50/50 (which the association threshold maps to "uncertain").
    counts: dict = {}
    for name, entry in data.items():
        normalized = normalize_name(name)
        if not normalized or len(entry) < 3:
            continue
        genders = entry[2]  # e.g. ["m"], ["f"], ["m", "f"]
        male = "m" in genders
        female = "f" in genders
        if male and female:
            bucket = counts.setdefault(normalized, [0, 0])
            bucket[0] += 50
            bucket[1] += 50
        elif male:
            bucket = counts.setdefault(normalized, [0, 0])
            bucket[1] += 100
        elif female:
            bucket = counts.setdefault(normalized, [0, 0])
            bucket[0] += 100
    print(f"  Meiji Yasuda: {len(counts):,} unique kanji names")
    return counts


def _load_china_mps() -> dict:
    """
    Official Chinese name statistics from MPS annual name reports.

    The Ministry of Public Security publishes annual reports with the
    most popular newborn names by sex (state media: People's Daily,
    China News). Names are embedded as official statistics; synthetic
    counts preserve the reported sex association.
    """
    counts: dict = {}
    for label, names in CHINA_MPS_NAMES.items():
        for name in names:
            normalized = normalize_name(name)
            if not normalized:
                continue
            bucket = counts.setdefault(normalized, [0, 0])
            if label == "girl-associated":
                bucket[0] += 100
            else:
                bucket[1] += 100
    print(f"  China MPS reports: {len(counts):,} unique names")
    return counts


def _download_to(url: str, dest: str, timeout: int = 300) -> bool:
    """Download url to dest unless dest already exists and is non-empty."""
    if os.path.exists(dest) and os.path.getsize(dest) > 0:
        return True
    os.makedirs(os.path.dirname(dest) or ".", exist_ok=True)
    try:
        req = urllib.request.Request(url, headers=_UA)
        with urllib.request.urlopen(req, timeout=timeout) as resp, \
                open(dest, "wb") as out:
            out.write(resp.read())
        return True
    except Exception as e:
        print(f"WARNING: download failed for {url}: {e}")
        return False


def _load_spain_ine() -> dict:
    """Spain INE: given-name frequency by sex (Padron 2025, frequency >= 20).

    Returns {name: (girl_count, boy_count)}. Names are stored without
    accents in the official file ("MARIA", "JOSE"); compounds such as
    "JOSE ANTONIO" are single given-name entries in Spanish records.
    """
    try:
        import openpyxl
    except ImportError:
        print("WARNING: openpyxl not installed, skipping Spain INE source")
        return {}

    dest = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "data", "raw", "es_nombres_frecuentes.xlsx")
    if not _download_to(ES_NAMES_URL, dest):
        return {}

    counts: dict = {}
    try:
        wb = openpyxl.load_workbook(dest, read_only=True)
    except Exception as e:
        print(f"WARNING: cannot read Spain INE file: {e}")
        return {}

    for ws in wb.worksheets:
        title = (ws.title or "").lower()
        is_girl = "mujer" in title
        is_boy = "hombre" in title
        if not (is_girl or is_boy):
            continue
        header_seen = False
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            if not header_seen:
                if any(str(c).upper() == "NOMBRE" for c in row if c is not None):
                    header_seen = True
                continue
            name = row[1] if len(row) > 1 else None
            freq = row[2] if len(row) > 2 else None
            if name is None:
                continue
            try:
                freq = int(freq)
            except (TypeError, ValueError):
                continue
            norm = normalize_name(str(name))
            if not norm or freq <= 0:
                continue
            bucket = counts.setdefault(norm, [0, 0])
            if is_girl:
                bucket[0] += freq
            else:
                bucket[1] += freq
    wb.close()
    return counts


def _load_sweden_isof() -> dict:
    """Sweden isoF newborn top-name tables (official, real counts).

    Each year file has "Flickor YYYY" (girls) and "Pojkar YYYY" (boys)
    sheets with columns [rank 2024, rank 2023, name, count, variants];
    only the main spelling and its real count are used.
    """
    try:
        import openpyxl
    except ImportError:
        print("WARNING: openpyxl not installed, skipping Sweden isoF source")
        return {}

    counts: dict = {}
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "raw")
    for year, url in SE_ISOF_FILES:
        dest = os.path.join(base, f"isof_babynamn_{year}.xlsx")
        if not _download_to(url, dest):
            continue
        try:
            wb = openpyxl.load_workbook(dest, read_only=True)
        except Exception as e:
            print(f"WARNING: cannot read Sweden isoF {year}: {e}")
            continue
        for ws in wb.worksheets:
            title = (ws.title or "").lower()
            is_girl = "flick" in title  # sheets are "Flickor YYYY" / "Pojkar YYYY"
            is_boy = "pojk" in title
            if not (is_girl or is_boy):
                continue
            rows = ws.iter_rows(values_only=True)
            col_name = col_count = None
            for row in rows:
                if not row:
                    continue
                for j, c in enumerate(row):
                    cs = str(c).strip().upper() if c is not None else ""
                    if col_name is None and cs == "NAMN":
                        col_name = j
                    elif col_count is None and (cs.startswith("ANTAL") or "TOTALT" in cs):
                        col_count = j
                if col_name is not None and col_count is not None:
                    break
            if col_name is None or col_count is None:
                continue
            for row in rows:
                if len(row) <= max(col_name, col_count):
                    continue
                name = row[col_name]
                if name is None:
                    continue
                try:
                    antal = int(row[col_count])
                except (TypeError, ValueError):
                    continue
                norm = normalize_name(str(name))
                if not norm or antal <= 0:
                    continue
                bucket = counts.setdefault(norm, [0, 0])
                if is_girl:
                    bucket[0] += antal
                else:
                    bucket[1] += antal
        wb.close()
    return counts


def _load_poland_pesel() -> dict:
    """Poland: full PESEL register of living people's first names.

    Returns {name: (girl_count, boy_count)}. CSV columns are
    [first name, sex, occurrence count]; the sex column is redundant
    because male and female lists are separate files.
    """
    import csv

    counts: dict = {}
    base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "raw")
    for sex, url in PL_PESEL_FILES:
        dest = os.path.join(base, f"pl_pesel_{sex}.csv")
        if not _download_to(url, dest):
            continue
        try:
            with open(dest, encoding="utf-8-sig", newline="") as fh:
                reader = csv.reader(fh)
                header = next(reader, None)
                if not header:
                    continue
                for row in reader:
                    if len(row) < 3:
                        continue
                    raw = (row[0] or "").strip()
                    if not raw:
                        continue
                    try:
                        cnt = int(row[2].strip())
                    except (TypeError, ValueError):
                        continue
                    norm = normalize_name(raw)
                    if not norm or cnt <= 0:
                        continue
                    bucket = counts.setdefault(norm, [0, 0])
                    if sex == "female":
                        bucket[0] += cnt
                    else:
                        bucket[1] += cnt
        except Exception as e:
            print(f"WARNING: cannot read Poland PESEL {sex} file: {e}")
            continue
    return counts


def _weight_for(total: int) -> float:
    """Weight by frequency: names with 10k+ records get full weight."""
    w = min(1.0, max(0.05, (total + 1) ** 0.25 / 10.0))
    return round(w, 4)


def main():
    base = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(base, "data")

    sources = {
        "US_SSA": _load_ssa(os.path.join(data_dir, "ssa_alldata.txt"),
                             url=SSA_ZIP_URL),
        "US_CENSUS": _load_census_xlsx(
            os.path.join(data_dir, "census_2020_firstnames_sex.xlsx")),
        "JP_MEIJI": _load_meiji_yasuda(),
        "CN_MPS": _load_china_mps(),
    }
    sources["FR_INSEE"] = _load_insee()
    sources["ES_INE"] = _load_spain_ine()
    sources["SE_ISOFF"] = _load_sweden_isof()
    sources["PL_PESEL"] = _load_poland_pesel()

    for source, counts in sources.items():
        total_records = sum(g + b for g, b in counts.values())
        print(f"  {source}: {len(counts):,} unique names, "
              f"{total_records:,} recorded births")

    # Merge across sources.
    merged: dict = {}
    for source, counts in sources.items():
        for name, (girl, boy) in counts.items():
            bucket = merged.setdefault(name, [0, 0, set()])
            bucket[0] += girl
            bucket[1] += boy
            bucket[2].add(source)

    entries = []
    for name, (girl, boy, srcs) in merged.items():
        total = girl + boy
        if total < 3:
            continue
        ratio = girl / total
        if ratio >= ASSOC_THRESHOLD:
            label = "girl-associated"
        elif ratio <= 1.0 - ASSOC_THRESHOLD:
            label = "boy-associated"
        else:
            label = "uncertain"
        entries.append({
            "name": name,
            "label": label,
            "weight": _weight_for(total),
            "girl_count": girl,
            "boy_count": boy,
            "total_records": total,
            "countries": ",".join(sorted(srcs)),
        })

    # Partition by normalized name so a name never leaks across splits.
    splits = {"train": [], "validation": [], "test": []}
    for entry in entries:
        h = hashlib.md5(entry["name"].encode("utf-8")).digest()[0] % 100
        if h < TRAIN_FRACTION * 100:
            splits["train"].append(entry)
        elif h < (TRAIN_FRACTION + VAL_FRACTION) * 100:
            splits["validation"].append(entry)
        else:
            splits["test"].append(entry)

    out_paths = {
        "train": os.path.join(data_dir, "train.jsonl"),
        "validation": os.path.join(data_dir, "validation.jsonl"),
        "test": os.path.join(data_dir, "test.jsonl"),
    }
    for key, path in out_paths.items():
        with open(path, "w", encoding="utf-8") as f:
            for entry in splits[key]:
                f.write(json.dumps(entry, ensure_ascii=False) + "\n")

    with open(os.path.join(data_dir, "names.jsonl"), "w", encoding="utf-8") as f:
        for entry in entries:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")

    report = {
        "sources": {s: {"unique_names": len(c)}
                    for s, c in sources.items()},
        "unique_names_total": len(entries),
        "total_recorded_births": sum(e["total_records"] for e in entries),
        "label_counts": {
            lbl: sum(1 for e in entries if e["label"] == lbl)
            for lbl in ("girl-associated", "boy-associated", "uncertain")
        },
        "split_sizes": {k: len(v) for k, v in splits.items()},
        "association_threshold": ASSOC_THRESHOLD,
    }
    with open(os.path.join(data_dir, "dataset_report.json"), "w",
              encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)

    print("\nDataset statistics")
    print("------------------")
    print(f"Total examples (unique names): {len(entries):,}")
    print(f"Total recorded births covered: {report['total_recorded_births']:,}")
    for lbl, n in report["label_counts"].items():
        print(f"{lbl}: {n:,}")
    print(f"Train: {len(splits['train']):,}  "
          f"Validation: {len(splits['validation']):,}  "
          f"Test: {len(splits['test']):,}")
    print(f"Report saved to {os.path.join(data_dir, 'dataset_report.json')}")


if __name__ == "__main__":
    main()